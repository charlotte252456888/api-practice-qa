"""Pytest hooks for collecting AI test results and exporting Excel reports.

這個檔案負責「報表輸出」：
- ai_report_recorder fixture 讓測試案例回報 prompt、latency、AI 回覆。
- pytest_runtest_makereport hook 取得每個測試最終 pass/fail 結果。
- pytest_sessionfinish hook 在全部測試結束後產出 AI_API_Performance_Report.xlsx。
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPORT_FILENAME = "AI_API_Performance_Report.xlsx"
REPORT_HEADERS = [
    "測試編號",
    "測試問題",
    "測試結果",
    "回應時間 (秒)",
    "答案字數",
    "回覆內容",
]


@dataclass
class AIReportRow:
    """Excel 報表中的一列資料。"""

    nodeid: str
    case_id: str
    prompt: str
    status: str
    latency_seconds: float
    answer_length: int
    reply_text: str


# 用 dict 暫存每個 test node 的報表資料。
# key 使用 pytest nodeid，能精準對應到每個測試函式。
AI_REPORT_ROWS: dict[str, AIReportRow] = {}


@pytest.fixture
def ai_report_recorder(request: pytest.FixtureRequest):
    """提供給測試案例呼叫，用來記錄 AI 回應資料。"""

    def record(
        *,
        case_id: str,
        prompt: str,
        latency_seconds: float,
        answer_length: int,
        reply_text: str,
    ) -> None:
        # 先預設為 Success；真正 pass/fail 會在 pytest_runtest_makereport 更新。
        AI_REPORT_ROWS[request.node.nodeid] = AIReportRow(
            nodeid=request.node.nodeid,
            case_id=case_id,
            prompt=prompt,
            status="Success",
            latency_seconds=latency_seconds,
            answer_length=answer_length,
            reply_text=reply_text,
        )

    return record


@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item: pytest.Item, call: pytest.CallInfo[Any]):
    """在每個測試結束後，取得 pytest 判定的最終結果。

    測試可能在 assert 才失敗，所以我們不能只在 AI 回應成功時就判定 Success。
    這個 hook 會把 call phase 的結果同步回 AI_REPORT_ROWS。
    """

    outcome = yield
    report = outcome.get_result()

    if report.when != "call":
        return

    row = AI_REPORT_ROWS.get(item.nodeid)
    if row is None:
        # 如果測試在 fixture 階段就 skip/fail，仍保留一列方便排查。
        AI_REPORT_ROWS[item.nodeid] = AIReportRow(
            nodeid=item.nodeid,
            case_id=item.name,
            prompt="",
            status=report.outcome.title(),
            latency_seconds=0.0,
            answer_length=0,
            reply_text=str(report.longrepr or ""),
        )
        return

    if report.passed:
        row.status = "Success"
    elif report.skipped:
        row.status = "Skipped"
    else:
        row.status = "Failed"
        # 測試失敗時，把 pytest 的錯誤摘要附在回覆欄後方，方便下載 Excel 排查。
        failure_text = str(report.longrepr or "")
        if failure_text and failure_text not in row.reply_text:
            row.reply_text = f"{row.reply_text}\n\nPytest failure:\n{failure_text}"


def pytest_sessionfinish(session: pytest.Session, exitstatus: int) -> None:
    """pytest 全部測試結束後，輸出 Excel 報表。"""

    workbook_path = Path.cwd() / REPORT_FILENAME
    _write_excel_report(workbook_path, AI_REPORT_ROWS.values())
    print(f"\nExcel report exported: {workbook_path}")


def _write_excel_report(workbook_path: Path, rows: Any) -> None:
    """沿用 main.py PoC 的 openpyxl 美化與 BarChart 報表邏輯。"""

    wb = Workbook()
    ws = wb.active
    ws.title = "API效能測試"
    ws.views.sheetView[0].showGridLines = True

    ws.append(REPORT_HEADERS)

    # 依測試編號排序，讓報表順序穩定、好讀。
    sorted_rows = sorted(rows, key=lambda row: row.case_id)
    for row in sorted_rows:
        ws.append(
            [
                row.case_id,
                row.prompt,
                row.status,
                row.latency_seconds,
                row.answer_length,
                row.reply_text,
            ]
        )

    if sorted_rows:
        avg_row_idx = len(sorted_rows) + 2
        ws.cell(row=avg_row_idx, column=3, value="平均回應時間").font = Font(
            name="Aptos", size=11, bold=True
        )
        ws.cell(
            row=avg_row_idx,
            column=4,
            value=f"=AVERAGE(D2:D{len(sorted_rows) + 1})",
        ).font = Font(name="Aptos", size=11, bold=True)
        ws.cell(row=avg_row_idx, column=4).number_format = "0.000"

    _style_worksheet(ws)

    if sorted_rows:
        _add_latency_chart(ws, len(sorted_rows))

    wb.save(workbook_path)


def _style_worksheet(ws) -> None:
    """設定標頭、邊框、對齊和自動欄寬。"""

    header_fill = PatternFill(
        start_color="2C3E50", end_color="2C3E50", fill_type="solid"
    )
    header_font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="BDC3C7"),
        right=Side(style="thin", color="BDC3C7"),
        top=Side(style="thin", color="BDC3C7"),
        bottom=Side(style="thin", color="BDC3C7"),
    )

    for col_num in range(1, len(REPORT_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            if cell.column in [1, 3]:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif cell.column in [4, 5]:
                cell.alignment = Alignment(horizontal="right", vertical="top")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 依內容自動調整欄寬；回覆內容可能很長，所以第 6 欄限制上限避免過寬。
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        if column_cells[0].column == 6:
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 20), 80)
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)


def _add_latency_chart(ws, result_count: int) -> None:
    """新增回應時間直條圖，方便快速比較各測項 latency。"""

    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = "AI API 回應時間比較 (Latency Baseline)"
    chart.y_axis.title = "回應時間 (秒)"
    chart.x_axis.title = "測試編號"

    chart_data = Reference(ws, min_col=4, min_row=1, max_row=result_count + 1)
    categories = Reference(ws, min_col=1, min_row=2, max_row=result_count + 1)
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend = None

    ws.add_chart(chart, "H2")
