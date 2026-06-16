import os
import time
import httpx  # 💡 確保最上方有 import httpx
from groq import Groq
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from dotenv import load_dotenv

# 1. 核心大絕招：建立一個「完全不驗證 SSL」的客製化連線通道
custom_http_client = httpx.Client(verify=False, timeout=15.0)

# 先載入同資料夾內的 .env，這樣就能把 API key 放在本機設定檔裡管理。
load_dotenv()

# 從環境變數讀取 API Key；如果沒有設定，就直接報錯，避免把金鑰寫死在程式裡。
API_KEY = os.environ.get("GROQ_API_KEY")
if not API_KEY:
    raise RuntimeError("請先設定環境變數 GROQ_API_KEY")
# 建立 Groq client，之後所有聊天請求都會透過這個物件送出。
client = Groq(
    api_key=API_KEY,
    http_client=custom_http_client  # 💡 就是這一行，強行突破防火牆！
)

# 測試問題清單：把多筆題目集中管理，方便一次批次測試。
test_prompts = [
    "請用一句話解釋什麼是自動化測試？",
    "什麼是 Python 的裝飾器 (Decorator)？請簡短說明。",
    "請寫出一個 Python 快速排序 (Quick Sort) 的範例程式碼。",
    "台灣最高的山是什麼山？高度是多少？",
    "請解釋軟體測試中的白箱測試與黑箱測試有何不同？",
    "如何優化網頁自動化測試中 time.sleep() 造成的硬等待問題？",
    "請列出 5 個現代 DevOps 常用的工具，並用一句話介紹其用途。",
    "什麼是 RESTful API？它的核心特點是什麼？",
    "請用中文寫一首關於『寫程式抓 Bug』的四句幽默短詩。",
    "如果自動化測試腳本在大規模執行時突然中斷，你身為 QA 工程師會從哪些方向排查？"
]

test_results = []
print("🚀 開始執行【Groq 極速版】AI API 自動化效能測試...")

for index, prompt in enumerate(test_prompts, start=1):
    print(f"[{index}/10] 正在測試問題: '{prompt[:15]}...'")
    start_time = time.time()
    try:
        # 呼叫 Groq 目前可用的模型。
        completion = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": prompt}],
            timeout=15.0  # 設定 15 秒超時控制
        )
        end_time = time.time()
        response_time = round(end_time - start_time, 3)
        reply_text = completion.choices[0].message.content
        answer_length = len(reply_text)
        status = "Success"
        print(f"  └ ✅ 測試成功！耗時: {response_time} 秒")
    except Exception as e:
        end_time = time.time()
        response_time = round(end_time - start_time, 3)
        reply_text = f"異常原因: {str(e)}"
        answer_length = 0
        status = "Failed"
        print(f"  └ ❌ 測試失敗：{e}")

    test_results.append([f"TC-{index:03d}", prompt, status, response_time, answer_length, reply_text])
    time.sleep(1)  # Groq 很快，冷卻時間 1 秒即可

print("\n📊 測試執行完畢，正在建構專業級 Excel 品質報告與圖表...")

# 以下為 OpenPyXL 自動繪圖邏輯（維持不變）
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "API效能測試"
ws.views.sheetView[0].showGridLines = True

headers = ["測試編號", "測試問題", "測試結果", "回應時間 (秒)", "答案字數", "AI 回覆內容"]
ws.append(headers)
for row in test_results:
    ws.append(row)

avg_row_idx = len(test_results) + 2
ws.cell(row=avg_row_idx, column=3, value="平均回應時間").font = Font(name="Aptos", size=11, bold=True)
ws.cell(row=avg_row_idx, column=4, value=f"=AVERAGE(D2:D{len(test_results)+1})").font = Font(name="Aptos", size=11, bold=True)
ws.cell(row=avg_row_idx, column=4).number_format = "0.000"

header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
header_font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
thin_border = Border(left=Side(style='thin', color='BDC3C7'), right=Side(style='thin', color='BDC3C7'),
                     top=Side(style='thin', color='BDC3C7'), bottom=Side(style='thin', color='BDC3C7'))

for col_num in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for r_idx in range(2, len(test_results) + 2):
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.border = thin_border
        if c_idx in [1, 3]: cell.alignment = Alignment(horizontal="center")
        elif c_idx in [4, 5]: cell.alignment = Alignment(horizontal="right")

chart = BarChart()
chart.type = "col"
chart.style = 11
chart.title = "Groq API 各測項回應時間 (Latency Baseline)"
chart.y_axis.title = "回應時間 (秒)"
chart.x_axis.title = "測試編號"

chart_data = Reference(ws, min_col=4, min_row=1, max_row=len(test_results) + 1)
cats = Reference(ws, min_col=1, min_row=2, max_row=len(test_results) + 1)
chart.add_data(chart_data, titles_from_data=True)
chart.set_categories(cats)
chart.legend = None

ws.add_chart(chart, "H2")

for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = openpyxl.utils.get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

report_filename = "AI_API_Performance_Report.xlsx"
wb.save(report_filename)
print(f"🎉 任務達成！品質報告已更新：{report_filename}")